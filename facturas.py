from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

# Leer el archivo Excel y cargarlo en un DataFrame
df_filtrado = pd.read_excel('datos_filtrados.xlsx')

def separar_bloques(df):

  # Obtener los indices de aquellas celdas que no estan vacias en 'GRUPO'
  indices = df[df['GRUPO'].notna()].index

  bloques = []

  ''' SEPARAR LOS BLOQUES '''
  for i in range(len(indices) - 1):

    inicio = indices[i] + 1
    fin = indices[i + 1] #- 1

    bloques.append(df.iloc[inicio:fin])
    print(inicio,fin)

  return bloques


# Obtenemos las separaciones
bloques = separar_bloques(df_filtrado)




def rellenar_fila(fila_a_insertar, info, ws, mayor="False"):
    # Iterar sobre las columnas y escribir los valores en las celdas correspondientes
    for idx, value in enumerate(info, start=1):
        ws.cell(row=fila_a_insertar, column=idx, value=value)



def aplicar_estilos(fila_a_formatear, ws, mayor="False"):
  # Establecer el ancho de la fila
  ws.row_dimensions[fila_a_formatear].height = 17.5

  # Establecer el estilo de fuente
  font_style = Font(name="Arial Narrow", size=13)  # Fuente Aptos Narrow de tamaño 13

  # Establecer el estilo de borde
  border_style = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

  # Crear un objeto Alignment con alineación centrada
  alignment = Alignment(horizontal='center', vertical='center')

  # Aplicar el estilo de fuente a todas las celdas de la fila
  for cell in ws[fila_a_formatear]:
      cell.font = font_style
      cell.border = border_style
      cell.alignment = alignment

  # Establecer el estilo de fuente en negrita

  cell = ws[f"A{fila_a_formatear}"]
  cell.font = Font(bold=True)
  if(mayor):
    cell = ws[f"K{fila_a_formatear}"]
    cell.font = Font(bold=True)
    cell = ws[f"H{fila_a_formatear}"]
    cell.number_format = '$#,##0.00'
    cell = ws[f"I{fila_a_formatear}"]
    cell.number_format = '$#,##0.00'
  else:
    cell = ws[f"I{fila_a_formatear}"]
    cell.font = Font(bold=True)
    cell = ws[f"G{fila_a_formatear}"]
    cell.number_format = '$#,##0.00'


def generar_info(bloque, ws, mayor="False"):
  guia = 1
  for indice in range(len(bloque)):
    registro = bloque.iloc[indice]
    if(mayor):
      info = [
          guia,
          registro['Tracking Number (HAWB)'],
          registro['TOTAL QTY OF ITEMS IN PARCEL'],
          "Paquete",
          registro['TOTAL QTY OF ITEMS IN PARCEL'],
          "Pz",
          registro['SHORT DESCRIPTION'],
          registro['TOTAL DECLARED VALUE'],
          (registro['TOTAL DECLARED VALUE'] * registro['TOTAL QTY OF ITEMS IN PARCEL']),
          "USA",
          "HOUND EXPRESS"
      ]
    else:
      info = [
          guia,
          registro['Tracking Number (HAWB)'],
          1,
          "Paquete",
          1,
          registro['SHORT DESCRIPTION'],
          registro['TOTAL DECLARED VALUE'],
          "USA",
          "HOUND EXPRESS"
      ]
    # print(info)
    # Insertar filas
    aumento = 6 if mayor else 8
    ws.insert_rows(guia + aumento, amount=1)
    rellenar_fila(guia + aumento, info, ws, mayor)
    aplicar_estilos(guia + aumento, ws, mayor)
    guia += 1

  if(mayor):
    ws[f'I{guia+aumento}'] = (bloque['TOTAL DECLARED VALUE'] * bloque['TOTAL QTY OF ITEMS IN PARCEL']).sum()
    ws[f'I{guia+aumento}'].number_format = '$#,##0.00'
    ws[f'E{guia+aumento}'] = bloque['TOTAL QTY OF ITEMS IN PARCEL'].sum()
  else:
    ws[f'G{guia+aumento}'] = bloque['TOTAL DECLARED VALUE'].sum()
    ws[f'G{guia+aumento}'].number_format = '$#,##0.00'



def generar_documento(titulo, bloque, mayor="False"):
  # Cargar el archivo Excel existente
  if(mayor):
    wb = load_workbook('plantilla_mayor.xlsx')
  else:
    wb = load_workbook('plantilla_menor.xlsx')
  # Seleccionar una hoja específica (o puedes usar wb.active para seleccionar la hoja activa)
  ws = wb.active

  # Eliminamos el registro de ejemplo
  ws.delete_rows(7 if mayor else 9)

  #
  generar_info(bloque, ws, mayor)


  # Guardar los cambios realizados en el archivo Excel
  wb.save(f'{titulo}.xlsx')


for indice, bloque in enumerate(bloques, start=0):
  if(indice == 0):
    generar_documento("FACTURA MENOR",bloque, False)
  else:
    generar_documento(f"FACTURA MAYOR{indice}",bloque, True)
